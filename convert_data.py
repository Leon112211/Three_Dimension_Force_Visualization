"""
convert_data.py
Reads Processed_Data.xlsx and outputs sensitivity_data.csv
for Processing to consume at runtime.

Output format (9 lines, 3 sensors x 3 rows each):
  # H2
  Sxx, Sxy, Sxz
  Syx, Syy, Syz
  Szx, Szy, Szz
  # H4
  ...
  # H6
  ...

Where S rows = Bx/By/Bz, cols = Fx/Fy/Fz (uT/N).
N/A or empty cells become 0.
"""

import os, sys

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
XLSX_PATH  = os.path.join(SCRIPT_DIR, "Processed_Data.xlsx")
CSV_PATH   = os.path.join(SCRIPT_DIR, "sensitivity_data.csv")

def to_float(v):
    if v is None or str(v).strip().upper() == "N/A" or str(v).strip() == "":
        return 0.0
    return float(v)

def build_matrix(data, col_offset):
    """Build 3x3 S matrix from 3 data rows starting at col_offset."""
    axis_map = {"X": 0, "Y": 1, "Z": 2}
    S = [[0.0]*3 for _ in range(3)]
    for i, row in enumerate(data):
        primary_axis = str(row[col_offset]).strip().upper()
        primary_val  = to_float(row[col_offset + 1])
        coup1_axis   = str(row[col_offset + 2]).strip().upper()
        coup1_val    = to_float(row[col_offset + 3])
        coup2_axis   = str(row[col_offset + 4]).strip().upper()
        coup2_val    = to_float(row[col_offset + 5])

        force_col = i  # row 0=Fx, 1=Fy, 2=Fz
        S[axis_map[primary_axis]][force_col] = primary_val
        S[axis_map[coup1_axis]][force_col]   = coup1_val
        S[axis_map[coup2_axis]][force_col]   = coup2_val
    return S

def main():
    if not os.path.exists(XLSX_PATH):
        print(f"ERROR: {XLSX_PATH} not found", file=sys.stderr)
        sys.exit(1)

    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[wb.sheetnames[0]]

    data = []
    for row in ws.iter_rows(min_row=3, max_row=5, values_only=True):
        data.append(list(row))

    sensor_names = ["H2", "H4", "H6"]
    offsets      = [0, 6, 12]

    lines = []
    for name, offset in zip(sensor_names, offsets):
        S = build_matrix(data, offset)
        lines.append(f"# {name}")
        for r in range(3):
            lines.append(f"{S[r][0]},{S[r][1]},{S[r][2]}")

    with open(CSV_PATH, "w", encoding="utf-8") as f:
        f.write("\n".join(lines) + "\n")

    print(f"OK: wrote {CSV_PATH}")

if __name__ == "__main__":
    main()
